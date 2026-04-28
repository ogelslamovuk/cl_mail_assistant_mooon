from __future__ import annotations

from dataclasses import dataclass
from pathlib import Path
from typing import Any

import openpyxl


@dataclass
class ExcelConfigStore:
    path: str

    def load(self) -> dict[str, Any]:
        workbook_path = Path(self.path)
        if not workbook_path.exists():
            raise FileNotFoundError(f"Excel config file not found: {workbook_path}")

        workbook = openpyxl.load_workbook(workbook_path, data_only=True)

        settings_sheet = self._resolve_required_sheet(workbook, ("SETTINGS", "НАСТРОЙКИ"))
        prompts_sheet = self._resolve_required_sheet(workbook, ("PROMPTS", "ПРОМПТЫ"))
        response_modes_sheet = self._resolve_required_sheet(workbook, ("RESPONSE_MODES", "РЕЖИМЫ ОТВЕТА"))
        text_policies_sheet = self._resolve_optional_sheet(workbook, ("TEXT_POLICIES", "ТЕКСТОВЫЕ ПОЛИТИКИ"))
        flags_sheet = self._resolve_optional_sheet(workbook, ("FLAGS", "ФЛАГИ"))

        settings = self._read_key_value_sheet(settings_sheet)
        prompt_rows = self._read_table_sheet(prompts_sheet)
        response_mode_rows = self._read_table_sheet(response_modes_sheet)
        text_policy_rows = self._read_table_sheet(text_policies_sheet) if text_policies_sheet is not None else []
        flag_rows = self._read_table_sheet(flags_sheet) if flags_sheet is not None else []

        prompts = self._collect_named_blocks(
            prompt_rows,
            required_ids={"system_prompt", "task_prompt", "static_support_context"},
        )
        text_policies = self._collect_named_blocks(text_policy_rows, required_ids=set())
        flags = self._collect_flags(flag_rows)
        allowed_response_modes = self._collect_response_modes(response_mode_rows)

        backend = str(settings.get("llm_backend", "") or "").strip()
        if backend not in {"ollama", "openai"}:
            raise RuntimeError("llm_backend must be either 'ollama' or 'openai'")

        prompt_version = str(settings.get("prompt_version", "") or "").strip() or self._derive_combined_version(prompt_rows)
        context_version = str(settings.get("context_version", "") or "").strip() or self._derive_combined_version(text_policy_rows)

        return {
            "source": str(workbook_path),
            "format": "excel",
            "backend": backend,
            "settings": settings,
            "prompts": prompts,
            "prompt_rows": prompt_rows,
            "text_policies": text_policies,
            "text_policy_rows": text_policy_rows,
            "flags": flags,
            "allowed_response_modes": allowed_response_modes,
            "prompt_version": prompt_version,
            "context_version": context_version,
        }

    def _resolve_required_sheet(self, workbook, aliases: tuple[str, ...]):
        for alias in aliases:
            if alias in workbook.sheetnames:
                return workbook[alias]
        joined = " / ".join(aliases)
        raise RuntimeError(f"Excel config must contain sheet: {joined}")

    def _resolve_optional_sheet(self, workbook, aliases: tuple[str, ...]):
        for alias in aliases:
            if alias in workbook.sheetnames:
                return workbook[alias]
        return None

    def _read_key_value_sheet(self, worksheet) -> dict[str, Any]:
        rows = list(worksheet.iter_rows(values_only=True))
        if not rows:
            return {}

        settings: dict[str, Any] = {}
        for index, row in enumerate(rows):
            if index == 0:
                continue
            key = row[0] if len(row) > 0 else None
            value = row[1] if len(row) > 1 else None
            value_type = row[2] if len(row) > 2 else "str"
            if key is None:
                continue
            key_text = str(key).strip()
            if not key_text:
                continue
            settings[key_text] = self._parse_typed_value(value=value, value_type=str(value_type or "str"))
        return settings

    def _read_table_sheet(self, worksheet) -> list[dict[str, Any]]:
        rows = list(worksheet.iter_rows(values_only=True))
        if not rows:
            return []

        header = [self._normalize_column_name(cell) for cell in rows[0]]
        records: list[dict[str, Any]] = []
        for raw_row in rows[1:]:
            record: dict[str, Any] = {}
            has_values = False
            for index, column_name in enumerate(header):
                if not column_name:
                    continue
                value = raw_row[index] if index < len(raw_row) else None
                if value is not None and str(value).strip() != "":
                    has_values = True
                record[column_name] = value
            if has_values:
                records.append(record)
        return records

    def _normalize_column_name(self, value: Any) -> str:
        raw = str(value).strip() if value is not None else ""
        mapping = {
            "идентификатор_блока": "block_id",
            "включено": "enabled",
            "версия": "version",
            "текст": "text",
            "описание": "description",
            "режим_ответа": "response_mode",
            "ключ": "key",
            "значение": "value",
            "тип": "type",
        }
        return mapping.get(raw, raw)

    def _collect_named_blocks(self, rows: list[dict[str, Any]], required_ids: set[str]) -> dict[str, dict[str, Any]]:
        blocks: dict[str, dict[str, Any]] = {}
        for row in rows:
            block_id = str(row.get("block_id", "") or "").strip()
            if not block_id:
                continue
            enabled = self._to_bool(row.get("enabled", 1))
            blocks[block_id] = {
                "block_id": block_id,
                "enabled": enabled,
                "version": str(row.get("version", "") or "").strip(),
                "text": str(row.get("text", "") or "").strip(),
                "description": str(row.get("description", "") or "").strip(),
            }

        missing = [block_id for block_id in required_ids if block_id not in blocks]
        if missing:
            joined = ", ".join(sorted(missing))
            raise RuntimeError(f"PROMPTS missing required block_id values: {joined}")
        return blocks

    def _collect_response_modes(self, rows: list[dict[str, Any]]) -> list[str]:
        modes: list[str] = []
        for row in rows:
            mode = str(row.get("response_mode", "") or "").strip()
            if not mode:
                continue
            enabled = self._to_bool(row.get("enabled", 1))
            if enabled:
                modes.append(mode)
        if not modes:
            raise RuntimeError("RESPONSE_MODES must contain at least one enabled response_mode")
        return modes

    def _collect_flags(self, rows: list[dict[str, Any]]) -> dict[str, Any]:
        flags: dict[str, Any] = {}
        for row in rows:
            key = str(row.get("key", "") or "").strip()
            if not key:
                continue
            value_type = str(row.get("type", "str") or "str")
            flags[key] = self._parse_typed_value(row.get("value"), value_type)
        return flags

    def _derive_combined_version(self, rows: list[dict[str, Any]]) -> str:
        versions: list[str] = []
        for row in rows:
            version = str(row.get("version", "") or "").strip()
            if version:
                versions.append(version)
        return "+".join(versions) if versions else "v1"

    def _parse_typed_value(self, value: Any, value_type: str) -> Any:
        normalized = value_type.strip().lower()
        if value is None:
            return None

        if normalized.startswith("bool"):
            return self._to_bool(value)
        if normalized == "int":
            return int(value)
        if normalized == "float":
            return float(value)
        if normalized in {"json", "dict", "list"}:
            import json
            return json.loads(str(value))
        if normalized in {"csv", "list_comma"}:
            return [part.strip() for part in str(value).split(",") if part.strip()]
        if normalized in {"pipe_list", "list_pipe"}:
            return [part.strip() for part in str(value).split("|") if part.strip()]
        return str(value)

    def _to_bool(self, value: Any) -> bool:
        if isinstance(value, bool):
            return value
        text = str(value).strip().lower()
        return text in {"1", "true", "yes", "y", "on"}
