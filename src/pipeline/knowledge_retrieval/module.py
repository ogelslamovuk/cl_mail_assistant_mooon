from __future__ import annotations

import re
from pathlib import Path
from typing import Any

from openpyxl import load_workbook

from src.pipeline.early_classification import early_classification
from src.pipeline.operator_flow import load_dossier, resolve_uid, save_dossier
from src.shared.common.paths import resolve_project_path
from src.shared.contracts.module_contract import ModuleResult
from src.shared.models.pipeline_context import PipelineContext



class KnowledgeRetrievalModule:
    name = "knowledge_retrieval"

    def __init__(
        self,
        dossier_path: str,
        knowledge_base_path: str = "config/knowledge_base/knowledge_base.xlsx",
    ) -> None:
        self.dossier_path = dossier_path
        self.knowledge_base_path = knowledge_base_path

    def run(self, context: PipelineContext) -> ModuleResult:
        try:
            dossier_path = resolve_project_path(self.dossier_path)
            payload = load_dossier(dossier_path)
            classification = early_classification(payload)
            if classification and classification.get("should_run_customer_kb") is False:
                reason = str(classification.get("classification_reason") or "early classification skipped customer KB")
                module_payload = _skipped_payload(payload, reason)
                payload.setdefault("modules", {})[self.name] = module_payload
                save_dossier(dossier_path, payload)
                context.artifacts.setdefault(self.name, []).append(str(dossier_path))
                return ModuleResult(
                    context=context,
                    status="ok",
                    notes=[f"knowledge_retrieval skipped uid={module_payload['uid']} reason={reason}"],
                    artifact_refs=[str(dossier_path)],
                    metrics=module_payload,
                )
            kb_path = resolve_project_path(self.knowledge_base_path)
            settings, rows, routing_rows = _load_knowledge_base(kb_path)
            query = _build_query(payload)
            matched = _score_rows(rows, query, settings)
            routing_matches = _score_routing_rows(routing_rows, query, settings)
            status = "found" if matched or routing_matches else "empty"
            module_payload = {
                "uid": resolve_uid(payload),
                "status": status,
                "error": "",
                "knowledge_base_path": str(kb_path),
                "query": {
                    "text": query["text"][:2000],
                    "response_mode": query["response_mode"],
                    "entities": query["entities"],
                },
                "matched_items": matched,
                "routing_matches": routing_matches,
                "settings": settings,
            }
            payload.setdefault("modules", {})[self.name] = module_payload
            save_dossier(dossier_path, payload)
        except Exception as exc:
            return ModuleResult(context=context, status="error", notes=[f"knowledge_retrieval failed: {exc}"])

        context.artifacts.setdefault(self.name, []).append(str(dossier_path))
        return ModuleResult(
            context=context,
            status="ok",
            notes=[
                f"knowledge_retrieval {status} uid={module_payload['uid']} items={len(matched)} routes={len(routing_matches)}"
            ],
            artifact_refs=[str(dossier_path)],
            metrics=module_payload,
        )


def _skipped_payload(payload: dict[str, Any], reason: str) -> dict[str, Any]:
    classification = early_classification(payload)
    return {
        "uid": resolve_uid(payload),
        "status": "skipped",
        "error": "",
        "reason": reason,
        "classification": str(classification.get("classification") or ""),
        "query": {
            "text": "",
            "response_mode": str(classification.get("response_mode_hint") or ""),
            "entities": [],
        },
        "matched_items": [],
        "routing_matches": [],
        "settings": {},
    }


def _load_knowledge_base(path: Path) -> tuple[dict[str, Any], list[dict[str, Any]], list[dict[str, Any]]]:
    if not path.exists():
        raise FileNotFoundError(f"Knowledge base not found: {path}")

    workbook = load_workbook(path, data_only=True)
    settings = {
        "min_score": 2,
        "max_results": 3,
        "max_routing_results": 2,
        "enabled_types": ["rule", "faq", "operator_instruction", "response_template", "regulation"],
    }
    if "KB_SETTINGS" in workbook.sheetnames:
        for row in workbook["KB_SETTINGS"].iter_rows(min_row=2, values_only=True):
            key = str(row[0] or "").strip()
            value = str(row[1] or "").strip()
            if not key:
                continue
            if key in {"min_score", "max_results", "max_routing_results"} and value.isdigit():
                settings[key] = int(value)
            elif key == "enabled_types":
                settings["enabled_types"] = [part.strip() for part in value.split(";") if part.strip()]
            else:
                settings[key] = value

    if "KB_ITEMS" not in workbook.sheetnames:
        raise ValueError("Knowledge base is missing KB_ITEMS sheet")

    sheet = workbook["KB_ITEMS"]
    headers = [str(cell.value or "").strip() for cell in sheet[1]]
    rows: list[dict[str, Any]] = []
    for values in sheet.iter_rows(min_row=2, values_only=True):
        item = {headers[index]: values[index] if index < len(values) else "" for index in range(len(headers))}
        if str(item.get("enabled") or "").strip().lower() not in {"yes", "true", "1"}:
            continue
        if str(item.get("type") or "").strip() not in set(settings["enabled_types"]):
            continue
        rows.append(item)

    routing_rows = _load_routing_rows(workbook)
    return settings, rows, routing_rows


def _load_routing_rows(workbook: Any) -> list[dict[str, Any]]:
    rows: list[dict[str, Any]] = []
    if "KB_ROUTING" in workbook.sheetnames:
        sheet = workbook["KB_ROUTING"]
        headers = [str(cell.value or "").strip() for cell in sheet[1]]
        for values in sheet.iter_rows(min_row=2, values_only=True):
            item = {headers[index]: values[index] if index < len(values) else "" for index in range(len(headers))}
            if str(item.get("enabled") or "").strip().lower() not in {"yes", "true", "1"}:
                continue
            rows.append(item)
    return _dedupe_routing_rows(rows)


def _dedupe_routing_rows(rows: list[dict[str, Any]]) -> list[dict[str, Any]]:
    result: list[dict[str, Any]] = []
    seen: set[str] = set()
    for row in rows:
        key = str(row.get("id") or row.get("email") or "").strip().casefold()
        if not key or key in seen:
            continue
        seen.add(key)
        result.append(row)
    return result


def _build_query(payload: dict[str, Any]) -> dict[str, Any]:
    headers = payload.get("headers") or {}
    modules = payload.get("modules") or {}
    llm = (modules.get("llm_understanding") or {}).get("structured_output") or {}
    entities = []
    for item in llm.get("entities") or []:
        if isinstance(item, dict):
            entities.append(str(item.get("value") or ""))
    text_parts = [
        headers.get("subject"),
        payload.get("body_text"),
        llm.get("summary"),
        llm.get("topic"),
        llm.get("customer_need"),
        " ".join(entities),
    ]
    return {
        "text": " ".join(str(part or "") for part in text_parts),
        "response_mode": str(llm.get("response_mode") or ""),
        "entities": [item for item in entities if item],
    }


def _score_rows(rows: list[dict[str, Any]], query: dict[str, Any], settings: dict[str, Any]) -> list[dict[str, Any]]:
    query_text = _normalize(query["text"])
    query_tokens = set(_tokens(query_text))
    response_mode = str(query.get("response_mode") or "")
    min_score = int(settings.get("min_score", 2) or 2)
    max_results = int(settings.get("max_results", 3) or 3)

    scored: list[tuple[float, dict[str, Any]]] = []
    for row in rows:
        score = 0.0
        keyword_hits: list[str] = []
        for column, weight in [
            ("topic_keywords", 4),
            ("customer_need_keywords", 5),
            ("entity_keywords", 3),
            ("title", 2),
            ("content", 1),
            ("operator_instruction", 1),
            ("template_hint", 1),
        ]:
            for keyword in _split_keywords(row.get(column)):
                normalized = _normalize(keyword)
                if not normalized:
                    continue
                if normalized in query_text:
                    score += weight
                    keyword_hits.append(keyword)
                    continue
                keyword_tokens = set(_tokens(normalized))
                if keyword_tokens and keyword_tokens <= query_tokens:
                    score += max(weight - 1, 1)
                    keyword_hits.append(keyword)

        allowed_modes = {mode.strip() for mode in str(row.get("response_modes") or "").split(";") if mode.strip()}
        if response_mode and allowed_modes and response_mode in allowed_modes:
            score += 2

        try:
            priority = float(row.get("priority") or 0)
        except Exception:
            priority = 0.0
        score += priority / 100.0

        if score >= min_score:
            scored.append((score, _compact_item(row, score, keyword_hits)))

    scored.sort(key=lambda pair: pair[0], reverse=True)
    return [item for _, item in scored[:max_results]]


def _score_routing_rows(rows: list[dict[str, Any]], query: dict[str, Any], settings: dict[str, Any]) -> list[dict[str, Any]]:
    query_text = _normalize(query["text"])
    query_tokens = set(_tokens(query_text))
    min_score = int(settings.get("min_score", 2) or 2)
    max_results = int(settings.get("max_routing_results", 2) or 2)

    scored: list[tuple[float, dict[str, Any]]] = []
    for row in rows:
        score = 0.0
        keyword_hits: list[str] = []
        for column, weight in [
            ("keywords", 5),
            ("description", 2),
            ("title", 2),
            ("department", 1),
            ("contact_name", 1),
            ("email", 1),
        ]:
            for keyword in _split_keywords(row.get(column)) if column == "keywords" else [str(row.get(column) or "")]:
                normalized = _normalize(keyword)
                if not normalized:
                    continue
                if normalized in query_text:
                    score += weight
                    keyword_hits.append(keyword)
                    continue
                keyword_tokens = set(_tokens(normalized))
                if keyword_tokens and keyword_tokens <= query_tokens:
                    score += max(weight - 1, 1)
                    keyword_hits.append(keyword)

        try:
            priority = float(row.get("priority") or 0)
        except Exception:
            priority = 0.0
        score += priority / 100.0

        if score >= min_score:
            scored.append((score, _compact_routing_item(row, score, keyword_hits)))

    scored.sort(key=lambda pair: pair[0], reverse=True)
    return [item for _, item in scored[:max_results]]


def _compact_item(row: dict[str, Any], score: float, keyword_hits: list[str]) -> dict[str, Any]:
    return {
        "id": str(row.get("id") or "").strip(),
        "category": str(row.get("category") or "").strip(),
        "type": str(row.get("type") or "").strip(),
        "title": str(row.get("title") or "").strip(),
        "score": round(score, 2),
        "content": str(row.get("content") or "").strip(),
        "operator_instruction": str(row.get("operator_instruction") or "").strip(),
        "template_hint": str(row.get("template_hint") or "").strip(),
        "source": str(row.get("source") or "").strip(),
        "source_ref": str(row.get("source_ref") or "").strip(),
        "matched_keywords": keyword_hits[:8],
    }


def _compact_routing_item(row: dict[str, Any], score: float, keyword_hits: list[str]) -> dict[str, Any]:
    return {
        "id": str(row.get("id") or "").strip(),
        "department": str(row.get("department") or "").strip(),
        "contact_name": str(row.get("contact_name") or "").strip(),
        "email": str(row.get("email") or "").strip(),
        "title": str(row.get("title") or "").strip(),
        "description": str(row.get("description") or "").strip(),
        "score": round(score, 2),
        "matched_keywords": keyword_hits[:8],
    }


def _split_keywords(value: Any) -> list[str]:
    return [part.strip() for part in str(value or "").split(";") if part and part.strip()]


def _normalize(value: str) -> str:
    return " ".join(str(value or "").casefold().replace("ё", "е").split())


def _tokens(value: str) -> list[str]:
    return re.findall(r"[0-9A-Za-zА-Яа-яЁё_.@-]+", value.casefold().replace("ё", "е"))
