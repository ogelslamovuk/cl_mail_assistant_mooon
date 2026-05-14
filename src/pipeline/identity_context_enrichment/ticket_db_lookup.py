from __future__ import annotations

import re
from dataclasses import dataclass, field
from pathlib import Path
from typing import Any, Dict, List, Optional, Tuple

try:
    import pymysql
except Exception:  # pragma: no cover
    pymysql = None

try:
    from openpyxl import load_workbook
except Exception:  # pragma: no cover
    load_workbook = None

try:
    from rapidfuzz import fuzz
    from rapidfuzz.distance import JaroWinkler
except Exception:  # pragma: no cover
    fuzz = None
    JaroWinkler = None

from src.layers.config.local_yaml_config_store import LocalYamlConfigStore
from src.shared.common.paths import project_root


ANCHOR_LEN_1 = 6
ANCHOR_LEN_2 = 4
MAX_CANDIDATES = 200
RESCUE_LIMIT = 250
TOP_N = 10
COMPACT_DROP_RE = re.compile(r"[._\-]+")

RAW_LOCAL_SQL = "LOWER(TRIM(SUBSTRING_INDEX(emailClient, '@', 1)))"
COMPACT_LOCAL_SQL = (
    "REPLACE(REPLACE(REPLACE(LOWER(TRIM(SUBSTRING_INDEX(emailClient, '@', 1))), '.', ''), '_', ''), '-', '')"
)


@dataclass
class TicketDbSettings:
    enabled: bool = False
    host: str = ""
    port: int = 3306
    user: str = ""
    password: str = ""
    view: str = "reports.searchtickets_report"
    lookback_days: int = 365
    connect_timeout: int = 10
    read_timeout: int = 15
    write_timeout: int = 15
    charset: str = "utf8mb4"
    source: str = "none"
    notes: list[str] = field(default_factory=list)


@dataclass
class TicketDbLookupResult:
    status: str
    rows: list[dict[str, Any]] = field(default_factory=list)
    rescue: list[dict[str, Any]] = field(default_factory=list)
    rescue_confident: bool = False
    rescue_reason: str = ""
    rescue_top1: float = 0.0
    rescue_gap: float = 0.0
    notes: list[str] = field(default_factory=list)
    settings_source: str = ""
    query_email: str = ""
    query_localpart: str = ""


class TicketDbLookupProvider:
    def __init__(self, settings: TicketDbSettings) -> None:
        self.settings = settings

    @classmethod
    def from_project_config(cls) -> "TicketDbLookupProvider":
        return cls(load_ticketdb_settings())

    def lookup_by_email(self, email: str) -> TicketDbLookupResult:
        local_in = localpart(email)
        if not self.settings.enabled:
            return TicketDbLookupResult(
                status="disabled",
                notes=list(self.settings.notes) or ["ticket_db provider disabled"],
                settings_source=self.settings.source,
                query_email=email,
                query_localpart=local_in,
            )

        if not local_in:
            return TicketDbLookupResult(
                status="empty",
                notes=["empty localpart after normalization"],
                settings_source=self.settings.source,
                query_email=email,
                query_localpart=local_in,
            )

        if pymysql is None:
            return TicketDbLookupResult(
                status="unavailable",
                notes=["ticket_db unavailable: pymysql is not installed"],
                settings_source=self.settings.source,
                query_email=email,
                query_localpart=local_in,
            )

        if fuzz is None or JaroWinkler is None:
            return TicketDbLookupResult(
                status="unavailable",
                notes=["ticket_db unavailable: rapidfuzz is not installed"],
                settings_source=self.settings.source,
                query_email=email,
                query_localpart=local_in,
            )

        try:
            conn = self._connect()
        except Exception as exc:
            return TicketDbLookupResult(
                status="error",
                notes=[f"ticket_db connect failed: {exc}"],
                settings_source=self.settings.source,
                query_email=email,
                query_localpart=local_in,
            )

        try:
            with conn.cursor() as cur:
                rows = pass1_exact_local(cur, local_in, self.settings)
                if rows:
                    return TicketDbLookupResult(
                        status="found_strict",
                        rows=rows,
                        notes=list(self.settings.notes),
                        settings_source=self.settings.source,
                        query_email=email,
                        query_localpart=local_in,
                    )

                rows = pass2_variant_local(cur, local_in, self.settings)
                if rows:
                    return TicketDbLookupResult(
                        status="found_strict",
                        rows=rows,
                        notes=list(self.settings.notes),
                        settings_source=self.settings.source,
                        query_email=email,
                        query_localpart=local_in,
                    )

                rescue = pass3_rescue_rapidfuzz(cur, local_in, self.settings, top_n=TOP_N)
                if rescue:
                    confident, reason, top1, gap = confident_auto_pick(rescue)
                    return TicketDbLookupResult(
                        status="rescue_candidates",
                        rescue=rescue,
                        rescue_confident=confident,
                        rescue_reason=reason,
                        rescue_top1=top1,
                        rescue_gap=gap,
                        notes=list(self.settings.notes),
                        settings_source=self.settings.source,
                        query_email=email,
                        query_localpart=local_in,
                    )

                return TicketDbLookupResult(
                    status="not_found",
                    notes=list(self.settings.notes),
                    settings_source=self.settings.source,
                    query_email=email,
                    query_localpart=local_in,
                )
        except Exception as exc:
            return TicketDbLookupResult(
                status="error",
                notes=[f"ticket_db query failed: {exc}"],
                settings_source=self.settings.source,
                query_email=email,
                query_localpart=local_in,
            )
        finally:
            try:
                conn.close()
            except Exception:
                pass

    def _connect(self):
        return pymysql.connect(
            host=self.settings.host,
            port=self.settings.port,
            user=self.settings.user,
            password=self.settings.password,
            charset=self.settings.charset,
            cursorclass=pymysql.cursors.DictCursor,
            connect_timeout=self.settings.connect_timeout,
            read_timeout=self.settings.read_timeout,
            write_timeout=self.settings.write_timeout,
        )


def load_ticketdb_settings() -> TicketDbSettings:
    yaml_settings = _load_from_local_yaml()
    if yaml_settings.enabled:
        return yaml_settings

    legacy_settings = _load_from_legacy_config_xlsx()
    if legacy_settings.enabled:
        return legacy_settings

    merged_notes = list(yaml_settings.notes) + list(legacy_settings.notes)
    return TicketDbSettings(enabled=False, source="none", notes=merged_notes or ["ticket_db config not found"])


def _load_from_local_yaml() -> TicketDbSettings:
    try:
        section = LocalYamlConfigStore().get_section("identity_context_enrichment")
    except FileNotFoundError:
        return TicketDbSettings(enabled=False, source="config.local.yaml", notes=["config.local.yaml not found"])
    except Exception as exc:
        return TicketDbSettings(enabled=False, source="config.local.yaml", notes=[f"identity_context_enrichment config unreadable: {exc}"])

    enabled = bool(section.get("ticketdb_enabled", False))
    password = str(section.get("ticketdb_password", "") or "")

    if not enabled:
        return TicketDbSettings(enabled=False, source="config.local.yaml", notes=["ticketdb_enabled is false or missing"])

    host = str(section.get("ticketdb_host", "") or "").strip()
    user = str(section.get("ticketdb_user", "") or "").strip()
    view = str(section.get("ticketdb_view", "reports.searchtickets_report") or "reports.searchtickets_report").strip()
    port = _safe_int(section.get("ticketdb_port"), 3306)
    lookback_days = _safe_int(section.get("ticketdb_lookback_days"), 365)
    connect_timeout = _safe_int(section.get("ticketdb_connect_timeout"), 10)
    read_timeout = _safe_int(section.get("ticketdb_read_timeout"), 15)
    write_timeout = _safe_int(section.get("ticketdb_write_timeout"), 15)

    notes: list[str] = []
    missing = [name for name, value in (("ticketdb_host", host), ("ticketdb_user", user), ("ticketdb_password", password)) if not value]
    if missing:
        return TicketDbSettings(enabled=False, source="config.local.yaml", notes=[f"missing TicketDB config keys: {', '.join(missing)}"])

    return TicketDbSettings(
        enabled=True,
        host=host,
        port=port,
        user=user,
        password=password,
        view=view,
        lookback_days=lookback_days,
        connect_timeout=connect_timeout,
        read_timeout=read_timeout,
        write_timeout=write_timeout,
        source="config.local.yaml",
        notes=notes,
    )


def _load_from_legacy_config_xlsx() -> TicketDbSettings:
    legacy_path = project_root() / "config.xlsx"
    if load_workbook is None:
        return TicketDbSettings(enabled=False, source="config.xlsx", notes=["openpyxl is not installed"])
    if not legacy_path.exists():
        return TicketDbSettings(enabled=False, source="config.xlsx", notes=["legacy config.xlsx not found"])

    try:
        wb = load_workbook(legacy_path, data_only=True)
    except Exception as exc:
        return TicketDbSettings(enabled=False, source="config.xlsx", notes=[f"cannot open legacy config.xlsx: {exc}"])

    if "SETTINGS" not in wb.sheetnames:
        return TicketDbSettings(enabled=False, source="config.xlsx", notes=["legacy config.xlsx has no SETTINGS sheet"])

    ws = wb["SETTINGS"]
    settings_map: dict[str, Any] = {}
    for row in range(2, ws.max_row + 1):
        key = ws.cell(row, 1).value
        if key is None:
            continue
        settings_map[str(key).strip()] = ws.cell(row, 2).value

    password = str(settings_map.get("db_password", "") or "")
    if not password:
        return TicketDbSettings(enabled=False, source="config.xlsx", notes=["legacy config.xlsx missing db_password"])

    return TicketDbSettings(
        enabled=True,
        host=str(settings_map.get("db_host", "soft.silverscreen.by") or "soft.silverscreen.by"),
        port=_safe_int(settings_map.get("db_port"), 3306),
        user=str(settings_map.get("db_user", "usr_ts") or "usr_ts"),
        password=password,
        view=str(settings_map.get("db_view", "reports.searchtickets_report") or "reports.searchtickets_report"),
        lookback_days=_safe_int(settings_map.get("db_lookup_lookback_days"), 365),
        connect_timeout=_safe_int(settings_map.get("db_connect_timeout"), 10),
        read_timeout=_safe_int(settings_map.get("db_read_timeout"), 15),
        write_timeout=_safe_int(settings_map.get("db_write_timeout"), 15),
        source="config.xlsx",
        notes=["using legacy config.xlsx fallback"],
    )


def _safe_int(value: Any, default: int) -> int:
    try:
        return int(value)
    except Exception:
        return default


def norm(s: str) -> str:
    return (s or "").strip().lower()


def normalize_localpart(lp: str) -> str:
    return "".join(norm(lp).split())


def compact_localpart(lp: str) -> str:
    return COMPACT_DROP_RE.sub("", normalize_localpart(lp))


def localpart(email: str) -> str:
    e = norm(email)
    lp = e.split("@", 1)[0] if "@" in e else e
    return normalize_localpart(lp)


def alpha_only(s: str) -> str:
    return re.sub(r"[^a-z]", "", norm(s))


def alpha_compact(s: str) -> str:
    return alpha_only(compact_localpart(s))


def digits_only(s: str) -> str:
    return "".join(ch for ch in norm(s) if ch.isdigit())


def chunk_tokens(s: str, n: int = 3) -> str:
    s = alpha_only(s)
    if not s:
        return ""
    if len(s) <= n:
        return s
    return " ".join(s[i : i + n] for i in range(0, len(s) - n + 1))


def swap_halves(s: str) -> str:
    s = alpha_only(s)
    if len(s) < 4:
        return s
    mid = len(s) // 2
    return s[mid:] + s[:mid]


def sorted_chars(s: str) -> str:
    return "".join(sorted(alpha_only(s)))


def pick_max_id_trading(ids) -> object:
    def _key(x):
        try:
            return int(x)
        except Exception:
            return str(x)

    cleaned = [i for i in ids if i is not None and str(i).strip() != ""]
    if not cleaned:
        return None
    return max(cleaned, key=_key)


def pick_best_row(rows: List[Dict[str, Any]]) -> Dict[str, Any]:
    def _key(r: Dict[str, Any]):
        t = r.get("ticket")
        try:
            t_num = int(t)
        except Exception:
            t_num = -1
        return (
            t_num,
            r.get("dateShow") or 0,
            r.get("timeShow") or 0,
            r.get("line") or 0,
            r.get("seat") or 0,
        )

    return max(rows, key=_key)


def is_edit_distance_leq_1(a: str, b: str) -> bool:
    if a == b:
        return True
    la, lb = len(a), len(b)
    if abs(la - lb) > 1:
        return False

    if la == lb:
        mism = []
        for i, (ca, cb) in enumerate(zip(a, b)):
            if ca != cb:
                mism.append(i)
                if len(mism) > 2:
                    return False
        if len(mism) == 1:
            return True
        if len(mism) == 2:
            i, j = mism
            return (j == i + 1) and (a[i] == b[j]) and (a[j] == b[i])
        return False

    if la > lb:
        a, b = b, a
        la, lb = lb, la

    i = j = 0
    used_skip = False
    while i < la and j < lb:
        if a[i] == b[j]:
            i += 1
            j += 1
            continue
        if used_skip:
            return False
        used_skip = True
        j += 1
    return True


def pass1_exact_local(cur, local_in: str, settings: TicketDbSettings) -> List[Dict[str, Any]]:
    cur.execute(
        f"""
        SELECT emailClient, idTrading, ticket, dateShow, timeShow, theater, event, auditorium, line, seat
        FROM {settings.view}
        WHERE {RAW_LOCAL_SQL} = %s
          AND dateShow >= (CURDATE() - INTERVAL {settings.lookback_days} DAY)
        ORDER BY dateShow DESC, timeShow DESC
        """,
        (local_in,),
    )
    rows = cur.fetchall()
    if not rows:
        return []

    ids = {r.get("idTrading") for r in rows}
    if len(ids) == 1:
        best = pick_best_row(rows)
        best["_match"] = "exact_localpart"
        return [best]

    chosen_id = pick_max_id_trading(ids)
    picked = [r for r in rows if r.get("idTrading") == chosen_id]
    best = pick_best_row(picked)
    best["_match"] = "exact_localpart_pick_max_idTrading"
    return [best]


def _like_prefix(base: str, wildcard_pos: Optional[int] = None) -> str:
    out: List[str] = []
    for i, ch in enumerate(base):
        if wildcard_pos is not None and i == wildcard_pos:
            out.append("_")
            continue
        if ch in ("\\", "%", "_"):
            out.append("\\" + ch)
        else:
            out.append(ch)
    return "".join(out) + "%"


def fetch_candidates(cur, local_in: str, like_pattern: str, settings: TicketDbSettings) -> List[Dict[str, Any]]:
    if not like_pattern:
        return []
    lo = max(len(local_in) - 1, 1)
    hi = len(local_in) + 1
    cur.execute(
        f"""
        SELECT emailClient, idTrading, ticket, dateShow, timeShow, theater, event, auditorium, line, seat
        FROM {settings.view}
        WHERE {RAW_LOCAL_SQL} LIKE %s ESCAPE '\\\\'
          AND CHAR_LENGTH({RAW_LOCAL_SQL}) BETWEEN %s AND %s
          AND dateShow >= (CURDATE() - INTERVAL {settings.lookback_days} DAY)
        ORDER BY dateShow DESC, timeShow DESC
        LIMIT %s
        """,
        (like_pattern, lo, hi, MAX_CANDIDATES),
    )
    return cur.fetchall()


def anchor_variants(local_in: str, anchor_len: int) -> List[str]:
    base = local_in[:anchor_len]
    out: List[str] = []
    seen = set()

    def _add(x: str):
        if x and x not in seen:
            seen.add(x)
            out.append(x)

    _add(base)
    chars = list(base)
    for i in range(len(base) - 1):
        tmp = chars.copy()
        tmp[i], tmp[i + 1] = tmp[i + 1], tmp[i]
        _add("".join(tmp))
    return out


def anchor_like_patterns(local_in: str, anchor_len: int) -> List[str]:
    base = local_in[:anchor_len]
    if not base:
        return []
    anchors = anchor_variants(local_in, anchor_len)
    pats: List[str] = []
    seen = set()

    def _add(p: str):
        if p and p not in seen:
            seen.add(p)
            pats.append(p)

    for a in anchors:
        _add(_like_prefix(a))
    for a in anchors:
        for i in range(len(a)):
            _add(_like_prefix(a, wildcard_pos=i))
    return pats


def fetch_candidates_transpose(cur, local_in: str, anchor_len: int, settings: TicketDbSettings) -> List[Dict[str, Any]]:
    for pat in anchor_like_patterns(local_in, anchor_len):
        rows = fetch_candidates(cur, local_in, pat, settings)
        if rows:
            return rows
    return []


def pass2_variant_local(cur, local_in: str, settings: TicketDbSettings) -> List[Dict[str, Any]]:
    candidates = fetch_candidates_transpose(cur, local_in, ANCHOR_LEN_1, settings)
    if not candidates and ANCHOR_LEN_2 < ANCHOR_LEN_1:
        candidates = fetch_candidates_transpose(cur, local_in, ANCHOR_LEN_2, settings)
    if not candidates:
        return []

    matches = []
    for r in candidates:
        cand_local = localpart(norm(r.get("emailClient") or ""))
        if is_edit_distance_leq_1(local_in, cand_local):
            matches.append(r)

    if not matches:
        return []

    ids = {r.get("idTrading") for r in matches}
    if len(ids) == 1:
        best = pick_best_row(matches)
        best["_match"] = "variant_localpart_distance<=1"
        return [best]

    chosen_id = pick_max_id_trading(ids)
    picked = [r for r in matches if r.get("idTrading") == chosen_id]
    best = pick_best_row(picked)
    best["_match"] = "variant_localpart_distance<=1_pick_max_idTrading"
    return [best]


def _dedup_rows(rows: List[Dict[str, Any]]) -> List[Dict[str, Any]]:
    dedup: Dict[tuple, Dict[str, Any]] = {}
    for row in rows:
        key = (
            str(row.get("emailClient") or "").lower(),
            str(row.get("idTrading") or ""),
            str(row.get("ticket") or ""),
        )
        dedup[key] = row
    return list(dedup.values())


def rescue_sql_patterns(local_in: str) -> List[str]:
    alpha = alpha_compact(local_in)
    parts: List[str] = []
    seen = set()

    def _add(x: str):
        if x and x not in seen:
            seen.add(x)
            parts.append(x)

    if len(alpha) >= 3:
        mid = len(alpha) // 2
        _add(alpha[:3])
        _add(alpha[-3:])
        _add(alpha[max(0, mid - 1) : max(0, mid - 1) + 3])

    if len(alpha) >= 4:
        _add(alpha[:4])
        _add(alpha[-4:])

    return [f"%{p}%" for p in parts if len(p) >= 3]


def fetch_rescue_candidates_primary(cur, local_in: str, settings: TicketDbSettings) -> List[Dict[str, Any]]:
    patterns = rescue_sql_patterns(local_in)
    if not patterns:
        return []

    clauses = " OR ".join([f"{COMPACT_LOCAL_SQL} LIKE %s ESCAPE '\\\\'" for _ in patterns])
    q_compact = compact_localpart(local_in)
    lo = max(len(q_compact) - 2, 1)
    hi = len(q_compact) + 3
    sql = f"""
        SELECT emailClient, idTrading, ticket, dateShow, timeShow, theater, event, auditorium, line, seat
        FROM {settings.view}
        WHERE ({clauses})
          AND CHAR_LENGTH({COMPACT_LOCAL_SQL}) BETWEEN %s AND %s
          AND dateShow >= (CURDATE() - INTERVAL {settings.lookback_days} DAY)
        ORDER BY dateShow DESC, timeShow DESC
        LIMIT %s
    """
    params = [*patterns, lo, hi, RESCUE_LIMIT]
    cur.execute(sql, params)
    return _dedup_rows(cur.fetchall())


def fetch_rescue_candidates_affix(cur, local_in: str, settings: TicketDbSettings) -> List[Dict[str, Any]]:
    q_compact = compact_localpart(local_in)
    q_alpha = alpha_compact(local_in)
    if len(q_alpha) < 4:
        return []

    prefix4 = q_alpha[:4]
    prefix5 = q_alpha[:5] if len(q_alpha) >= 5 else q_alpha[:4]
    suffix4 = q_alpha[-4:]

    clauses = [
        f"{COMPACT_LOCAL_SQL} LIKE %s ESCAPE '\\\\'",
        f"{COMPACT_LOCAL_SQL} LIKE %s ESCAPE '\\\\'",
        f"{COMPACT_LOCAL_SQL} LIKE %s ESCAPE '\\\\'",
    ]
    params = [
        f"{prefix5}%",
        f"{prefix4}%",
        f"%{suffix4}%",
    ]

    lo = max(len(q_compact) - 2, 1)
    hi = len(q_compact) + 8
    sql = f"""
        SELECT emailClient, idTrading, ticket, dateShow, timeShow, theater, event, auditorium, line, seat
        FROM {settings.view}
        WHERE ({' OR '.join(clauses)})
          AND CHAR_LENGTH({COMPACT_LOCAL_SQL}) BETWEEN %s AND %s
          AND dateShow >= (CURDATE() - INTERVAL {settings.lookback_days} DAY)
        ORDER BY dateShow DESC, timeShow DESC
        LIMIT %s
    """
    cur.execute(sql, [*params, lo, hi, RESCUE_LIMIT])
    return _dedup_rows(cur.fetchall())


def score_candidate(query_local: str, candidate_email: str) -> Dict[str, Any]:
    cand_local = localpart(candidate_email)
    q_raw = query_local
    c_raw = cand_local

    q_compact = compact_localpart(q_raw)
    c_compact = compact_localpart(c_raw)

    q_alpha = alpha_only(q_raw)
    c_alpha = alpha_only(c_raw)
    q_alpha_compact = alpha_compact(q_raw)
    c_alpha_compact = alpha_compact(c_raw)

    ratio = float(fuzz.ratio(q_raw, c_raw)) / 100.0
    partial = float(fuzz.partial_ratio(q_raw, c_raw)) / 100.0
    jw = float(JaroWinkler.normalized_similarity(q_raw, c_raw))

    compact_ratio = float(fuzz.ratio(q_compact, c_compact)) / 100.0 if q_compact and c_compact else 0.0
    compact_partial = float(fuzz.partial_ratio(q_compact, c_compact)) / 100.0 if q_compact and c_compact else 0.0
    compact_jw = float(JaroWinkler.normalized_similarity(q_compact, c_compact)) if q_compact and c_compact else 0.0

    if q_alpha and c_alpha:
        alpha_ratio = float(fuzz.ratio(q_alpha, c_alpha)) / 100.0
        alpha_partial = float(fuzz.partial_ratio(q_alpha, c_alpha)) / 100.0
        alpha_jw = float(JaroWinkler.normalized_similarity(q_alpha, c_alpha))
        chunk_ratio = float(fuzz.token_sort_ratio(chunk_tokens(q_alpha_compact), chunk_tokens(c_alpha_compact))) / 100.0
        bag_ratio = float(fuzz.ratio(sorted_chars(q_alpha_compact), sorted_chars(c_alpha_compact))) / 100.0
        swap_ratio = float(fuzz.ratio(swap_halves(q_alpha_compact), c_alpha_compact)) / 100.0
    else:
        alpha_ratio = alpha_partial = alpha_jw = chunk_ratio = bag_ratio = swap_ratio = 0.0

    q_digits = digits_only(q_raw)
    c_digits = digits_only(c_raw)
    digits_bonus = 0.04 if q_digits and q_digits == c_digits else 0.0
    digits_penalty = -0.03 if q_digits and c_digits and q_digits != c_digits else 0.0

    affix_bonus = 0.0
    if q_alpha_compact and c_alpha_compact:
        if c_alpha_compact.startswith(q_alpha_compact) or q_alpha_compact.startswith(c_alpha_compact):
            affix_bonus += 0.06
        elif c_alpha_compact.endswith(q_alpha_compact) or q_alpha_compact.endswith(c_alpha_compact):
            affix_bonus += 0.04

    score = (
        bag_ratio * 0.18
        + chunk_ratio * 0.18
        + compact_jw * 0.14
        + compact_ratio * 0.12
        + swap_ratio * 0.12
        + alpha_jw * 0.10
        + alpha_ratio * 0.06
        + compact_partial * 0.04
        + jw * 0.03
        + ratio * 0.01
        + partial * 0.01
        + digits_bonus
        + digits_penalty
        + affix_bonus
    )
    score = max(0.0, min(1.0, score))

    return {
        "query_local": q_raw,
        "candidate_local": c_raw,
        "candidate_email": candidate_email,
        "score": score,
        "ratio": ratio,
        "partial": partial,
        "jw": jw,
        "compact_ratio": compact_ratio,
        "compact_partial": compact_partial,
        "compact_jw": compact_jw,
        "alpha_ratio": alpha_ratio,
        "alpha_partial": alpha_partial,
        "alpha_jw": alpha_jw,
        "chunk_ratio": chunk_ratio,
        "bag_ratio": bag_ratio,
        "swap_ratio": swap_ratio,
        "digits_query": q_digits,
        "digits_candidate": c_digits,
        "affix_bonus": affix_bonus,
    }


def _collapse_ranked_groups(scored_rows: List[Dict[str, Any]], top_n: int) -> List[Dict[str, Any]]:
    grouped: Dict[str, List[Dict[str, Any]]] = {}
    ordered_keys: List[str] = []
    for row in scored_rows:
        id_trading = str(row.get("idTrading") or "").strip()
        if id_trading:
            group_key = f"id:{id_trading}"
        else:
            group_key = f"email:{str(row.get('emailClient') or '').strip().lower()}"
        if group_key not in grouped:
            grouped[group_key] = []
            ordered_keys.append(group_key)
        grouped[group_key].append(row)

    collapsed: List[Dict[str, Any]] = []
    for group_key in ordered_keys:
        rows = grouped[group_key]
        best = pick_best_row(rows)
        best["_rf_group_key"] = group_key
        best["_rf_group_size"] = len(rows)
        collapsed.append(best)

    collapsed.sort(
        key=lambda r: (
            float(r.get("_rf_score") or 0.0),
            r.get("dateShow") or 0,
            r.get("timeShow") or 0,
            r.get("ticket") or 0,
        ),
        reverse=True,
    )
    return collapsed[:top_n]


def rank_candidates(local_in: str, rows: List[Dict[str, Any]], source: str, top_n: int = TOP_N) -> List[Dict[str, Any]]:
    scored: List[Dict[str, Any]] = []
    for row in rows:
        s = score_candidate(local_in, str(row.get("emailClient") or ""))
        row_copy = dict(row)
        row_copy.update({f"_rf_{k}": v for k, v in s.items()})
        row_copy["_match"] = source
        scored.append(row_copy)

    scored.sort(
        key=lambda r: (
            float(r.get("_rf_score") or 0.0),
            r.get("dateShow") or 0,
            r.get("timeShow") or 0,
            r.get("ticket") or 0,
        ),
        reverse=True,
    )
    return _collapse_ranked_groups(scored, top_n)


def confident_auto_pick(rescue: List[Dict[str, Any]]) -> Tuple[bool, str, float, float]:
    if not rescue:
        return False, "none", 0.0, 0.0

    top1 = float(rescue[0].get("_rf_score") or 0.0)
    top2 = float(rescue[1].get("_rf_score") or 0.0) if len(rescue) > 1 else 0.0
    gap = top1 - top2

    bag = float(rescue[0].get("_rf_bag_ratio") or 0.0)
    chunk = float(rescue[0].get("_rf_chunk_ratio") or 0.0)
    swap = float(rescue[0].get("_rf_swap_ratio") or 0.0)
    compact_ratio = float(rescue[0].get("_rf_compact_ratio") or 0.0)
    compact_jw = float(rescue[0].get("_rf_compact_jw") or 0.0)
    affix_bonus = float(rescue[0].get("_rf_affix_bonus") or 0.0)

    if top1 >= 0.93 and gap >= 0.06:
        return True, "strong_score", top1, gap

    structural_hit = (
        bag >= 0.96
        and (chunk >= 0.78 or swap >= 0.88 or compact_ratio >= 0.90 or affix_bonus > 0.0)
        and compact_jw >= 0.70
    )
    if top1 >= 0.84 and gap >= 0.12 and structural_hit:
        return True, "structural", top1, gap

    return False, "manual_review", top1, gap


def pass3_rescue_rapidfuzz(cur, local_in: str, settings: TicketDbSettings, top_n: int = TOP_N) -> List[Dict[str, Any]]:
    primary = fetch_rescue_candidates_primary(cur, local_in, settings)
    ranked_primary = rank_candidates(local_in, primary, "rescue_rapidfuzz_primary", top_n=top_n) if primary else []
    confident_primary, _, _, _ = confident_auto_pick(ranked_primary)
    if confident_primary:
        return ranked_primary

    affix = fetch_rescue_candidates_affix(cur, local_in, settings)
    merged_by_key: Dict[tuple, Dict[str, Any]] = {}
    for row in primary + affix:
        key = (
            str(row.get("emailClient") or "").lower(),
            str(row.get("idTrading") or ""),
            str(row.get("ticket") or ""),
        )
        merged_by_key[key] = row

    merged_rows = list(merged_by_key.values())
    if not merged_rows:
        return []
    return rank_candidates(local_in, merged_rows, "rescue_rapidfuzz_v2", top_n=top_n)
